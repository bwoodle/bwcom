"""Workbook parsing for legacy race spreadsheets."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from legacy_races_pipeline.models import ParsedWorkbookRow


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _worksheet_has_custom_styles(worksheet: Worksheet) -> bool:
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if bool(cell.font and cell.font.strike):
                return True
            if cell.style_id != 0:
                return True
    return False


def parse_workbook(
    path: str | Path,
    sheet_name: str | None = None,
) -> list[ParsedWorkbookRow]:
    """Parse workbook rows with formatting metadata preserved when available."""
    workbook = load_workbook(filename=Path(path), data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    headers = [_normalize_header(cell.value) for cell in worksheet[1]]
    formatting_available = _worksheet_has_custom_styles(worksheet)
    parsed_rows: list[ParsedWorkbookRow] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        values: dict[str, object] = {}
        has_strikethrough = False

        for index, cell in enumerate(row):
            header = headers[index] if index < len(headers) else ""
            if not header or cell.value is None:
                continue
            values[header] = cell.value
            if formatting_available and bool(cell.font and cell.font.strike):
                has_strikethrough = True

        if not values:
            continue

        parsed_rows.append(
            ParsedWorkbookRow(
                row_id=f"{worksheet.title}!{row_number}",
                source_sheet=worksheet.title,
                values=values,
                formatting_available=formatting_available,
                has_strikethrough=has_strikethrough,
            )
        )

    return parsed_rows
